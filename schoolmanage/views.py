from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest
from django.template import loader
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist

import datetime
from .models import Eleve, Enseignant, Evaluation, Classe, Matiere, Classe_Matiere, Resultat, Note
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.core.files.storage import FileSystemStorage
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.db.models import Avg
# from .admin import predict_orientation
from orientation.models import Filiere, Filiere_Matiere


#OCR
from PIL import Image
from .OcrToTableTool import OcrToTableTool as ottt
from .TableExtractor import TableExtractor as te
from .TableLinesRemover import TableLinesRemover as tlr

import cv2
import matplotlib.pyplot as plt

import os
import openpyxl
import pytesseract

pytesseract.pytesseract.tesseract_cmd = "C:\\Program Files\\Tesseract-OCR\\tesseract.exe"


def predict_orientation(eleve):
    """
    Fonction pour prédire l'orientation d'un élève en fonction de ses notes dans les matières et des critères d'admission des filières.
    """
    # Récupérer les critères d'admission pour chaque filière
    filieres = Filiere.objects.prefetch_related('filiere_matiere_set').all()

    # Initialiser un dictionnaire pour stocker les notes moyennes de l'élève dans chaque matière
    moyennes_par_matiere = {}

    # Calculer les moyennes des notes de l'élève dans chaque matière
    for classe_matiere in eleve.classe.classe_matiere_set.all():
        #Récupération des résultats associés à cette classe_matiere
        resultats = Resultat.objects.filter(classematiere_id=classe_matiere)
        notes = Note.objects.filter(eleve_id=eleve, resultat_id__in=resultats)
        # notes = Note.objects.filter(eleve_id=eleve, resultat_id__classematiere=classe_matiere)
        if notes.exists():
            moyenne = sum(float(note.note) for note in notes) / len(notes)
            moyennes_par_matiere[classe_matiere.matiere_id] = moyenne

    # Prédire l'orientation de l'élève en fonction des critères d'admission pour chaque filière
    filiere_predite = None
    meilleur_score = 0
    for filiere in filieres:
        score_filiere = 0
        for filiere_matiere in filiere.filiere_matiere_set.all():
            if filiere_matiere.matiere_id in moyennes_par_matiere:
                if moyennes_par_matiere[filiere_matiere.matiere_id] >= filiere_matiere.note:
                    score_filiere += 1
        if score_filiere > meilleur_score:
            filiere_predite = filiere
            meilleur_score = score_filiere

    return filiere_predite

def predict_orientation_view(request, eleve_id):
    # Récupérer l'élève
    eleve = get_object_or_404(Eleve, pk=eleve_id)

    # Appeler la fonction de prédiction d'orientation
    filiere_predite = predict_orientation(eleve)

    # Mettre à jour la prédiction d'orientation de l'élève
    eleve.filiere = filiere_predite.name
    eleve.save()

    # Ajouter un message de succès
    messages.success(request, f"Prédiction d'orientation pour {eleve.name} mise à jour avec succès.")

    return redirect('admin:schoolmanage_eleve_changelist')

def home(request):
    context = {"message": "Hello World !",
               }

    return render(request,"schoolmanage/index.html", context)

def is_not_superuser(user):
    return user.is_superuser == False

def direction(request):
    
    return render(request,"schoolmanage/direction.html")

def parents(request):
    
    return render(request,"schoolmanage/parents.html")

#@permission_required('mangalib.view_book', raise_exception = True)
@user_passes_test(is_not_superuser)
@login_required
def enseignants(request):
    user = request.user

    classe_ids = Classe_Matiere.objects.filter(user=user).values_list('classe_id', flat=True)

    classes =  Classe.objects.filter(pk__in = classe_ids)
    
    context = {
        "classes": classes
    }
    return render(request,"schoolmanage/enseignants.html", context)

@user_passes_test(is_not_superuser)
@login_required
def show_matiere(request, classe_id):
    user = request.user

    request.session['classe_id'] = classe_id

    matiere_ids = Classe_Matiere.objects.filter(user=user, classe_id=classe_id).values_list('matiere_id', flat=True)

    classe = Classe.objects.get(pk = classe_id)

    matieres =  Matiere.objects.filter(pk__in = matiere_ids)
    
    context = {
        "classe" : classe,
        "matieres": matieres
    }
    return render(request, "schoolmanage/enseignants/classes/index.html", context)

@user_passes_test(is_not_superuser)
@login_required
def show_evaluation(request, matiere_id):
    user = request.user

    request.session['matiere_id'] = matiere_id

    classe_id = request.session['classe_id'] #request.session.get('matiere_id')

    matiere_ids = Classe_Matiere.objects.filter(user=user, classe_id=classe_id).values_list('matiere_id', flat=True)

    classe = Classe.objects.get(pk = classe_id)
    matiere = Matiere.objects.get(pk = matiere_id)

    evaluations =  Evaluation.objects.all()
    
    context = {
        "classe" : classe,
        "matiere": matiere,
        "evaluations": evaluations
    }
    return render(request, "schoolmanage/enseignants/evaluations/index.html", context)

@user_passes_test(is_not_superuser)
@login_required
def index_note(request, evaluation_id):
    user = request.user

    user = request.user

    matiere_id = request.session['matiere_id'] 
    classe_id = request.session['classe_id'] 
    request.session['evaluation_id'] = evaluation_id

    matiere_ids = Classe_Matiere.objects.filter(user=user, classe_id=classe_id).values_list('matiere_id', flat=True)

    classe = Classe.objects.get(pk = classe_id)
    matiere = Matiere.objects.get(pk = matiere_id)
    evaluation = Evaluation.objects.get(pk = evaluation_id)

    classematiere = Classe_Matiere.objects.get(classe_id=classe_id, matiere_id=matiere_id)

    try:
        resultat = Resultat.objects.get(classematiere_id=classematiere, evaluation_id=evaluation)
        eleves_notes = Note.objects.filter(resultat_id=resultat)
    except ObjectDoesNotExist:
        resultat = None  # Si aucun résultat n'est trouvé, définissez resultat sur None ou tout autre valeur par défaut.
        eleves_notes = [] 

    context = {
        "classe" : classe,
        "matiere": matiere,
        "evaluation": evaluation,
        "eleves_notes": eleves_notes
    }
    return render(request, "schoolmanage/enseignants/notes/index.html", context)


@user_passes_test(is_not_superuser)
@login_required
def edit_note(request, eleve_note_id):
    user = request.user

    eleve_note = Note.objects.get(pk = eleve_note_id)

    if request.method == 'POST':
        note = request.POST["note"]

        eleve_note.note = note

        eleve_note.save()

        return redirect("schoolmanage:index_note", evaluation_id = request.session['evaluation_id'] )

    matiere_id = request.session['matiere_id'] 
    classe_id = request.session['classe_id'] 
    evaluation_id = request.session['evaluation_id'] 

    matiere_ids = Classe_Matiere.objects.filter(user=user, classe_id=classe_id).values_list('matiere_id', flat=True)

    classe = Classe.objects.get(pk = classe_id)
    matiere = Matiere.objects.get(pk = matiere_id)
    evaluation = Evaluation.objects.get(pk = evaluation_id)
   
    context = {
        "classe" : classe,
        "matiere": matiere,
        "evaluation": evaluation,
        "eleve_note": eleve_note
    }
    return render(request, "schoolmanage/enseignants/notes/edit.html", context)

@user_passes_test(is_not_superuser)
@login_required
def add_note(request, evaluation_id):
    user = request.user

    matiere_id = request.session['matiere_id'] 
    classe_id = request.session['classe_id'] 
    request.session['evaluation_id'] = evaluation_id

    matiere_ids = Classe_Matiere.objects.filter(user=user, classe_id=classe_id).values_list('matiere_id', flat=True)

    classe = Classe.objects.get(pk = classe_id)
    matiere = Matiere.objects.get(pk = matiere_id)
    evaluation = Evaluation.objects.get(pk = evaluation_id)

    
    context = {
        "classe" : classe,
        "matiere": matiere,
        "evaluation": evaluation,
    }

    if request.method == 'POST' and request.FILES['file_note']:
        file_note = request.FILES['file_note']

        # Créer le dossier s'il n'existe pas déjà
        app_folder = os.path.dirname(__file__)
        folder_path = os.path.join(app_folder, 'files_notes')
        os.makedirs(folder_path, exist_ok=True)

        # Enregistrer le fichier dans le dossier files_notes avec le nom 1.png
        file_path = os.path.join(folder_path, '1.jpg')
        with open(file_path, 'wb') as destination:
            for chunk in file_note.chunks():
                destination.write(chunk)
    
        dossier = "ocr_slices"

        # Supprimer tous les fichiers du dossier ocr_slices
        current_directory = os.path.dirname(__file__)

        # Parcourir les fichiers dans le dossier
        for fichier in os.listdir(os.path.join(current_directory, dossier)):
            if os.path.isfile(os.path.join(current_directory, dossier, fichier)):
                # Construire le chemin absolu du fichier
                chemin_fichier = os.path.join(current_directory, dossier, fichier)
                # Supprimer le fichier
                os.remove(chemin_fichier)


        image_uploaded_path = os.path.join(os.path.dirname(__file__), 'files_notes', '1.jpg')

        # OCERISATION
        table_extractor = te(image_uploaded_path)
        perspective_corrected_image = table_extractor.execute()

        lines_remover = tlr(perspective_corrected_image)
        image_without_lines = lines_remover.execute()

        ocr_tool = ottt(image_without_lines, perspective_corrected_image)
        ocr_tool.execute()

        # GENERER LE FICHIER EXCEL
        dossier = "ocr_slices"
        element = []

        # Initialiser le compteur de fichiers
        nombre_fichiers = 0

        # Récupérer le répertoire contenant le fichier views.py
        current_directory = os.path.dirname(__file__)

        # Parcourir les fichiers dans le dossier
        for fichier in os.listdir(os.path.join(current_directory, dossier)):
            if os.path.isfile(os.path.join(current_directory, dossier, fichier)):
                img = cv2.imread(os.path.join(current_directory, dossier, "img_" + str(nombre_fichiers) + ".jpg")) 
                output = pytesseract.image_to_string(img, config='--psm 6 tessedit_char_whitelist=0123456789')
                output = output.strip()
                element.append(output)
                nombre_fichiers += 1
        
        excel_file_path = os.path.join(current_directory, "tableau.xlsx")
        # Vérifier si le FICHIER EXCEL EXISTE AVANT DE LE SUPPRIMER
        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)
            print("Le fichier Excel a été supprimé avec succès.")
        else:
            print("Le fichier Excel n'existe pas.")

        # CREER LE CLASSEUR EXCEL
        wb = openpyxl.Workbook()
        # Sélectionner la première feuille
        sheet = wb.active
        current_directory = os.path.dirname(__file__)
        
        # Définir les données à insérer dans le tableau
        data = element

        # Parcourir les éléments de la liste et les écrire dans le tableau Excel
        row = 1
        col = 1
        for item in data:
            sheet.cell(row=row, column=col, value=item)
            col += 1
            # Si nous avons atteint la troisième colonne, passer à la colonne suivante et réinitialiser la colonne à 1
            if col > 3:
                col = 1
                row += 1

        wb.save(os.path.join(current_directory, "tableau.xlsx"))
        print(element)
        
        #RECUPERER LES VALEURS DANS DES VARIABLES
        wb = openpyxl.load_workbook(os.path.join(current_directory, "tableau.xlsx"))
        feuille = wb.active

        colonne_note = []
        colonne_matricule = []
        colonne_noms = []

        for row in feuille.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                colonne_matricule.append(cell.value)
        print("MATRICULE " )
        print( colonne_matricule)
        for row in feuille.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                colonne_noms.append(cell.value)
        print("NOMS :" )
        print(colonne_noms)
        for row in feuille.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                colonne_note.append(cell.value)
        print("NOTE :" )
        print(colonne_note)


        classematiere = Classe_Matiere.objects.get(classe_id=classe_id, matiere_id=matiere_id)

        eleves = Eleve.objects.filter(classe = classe_id)

        
        resultat = Resultat.objects.create(classematiere_id = classematiere, evaluation_id = evaluation)

        i = 0
        for eleve in eleves:
            if eleve.name == colonne_noms[i]:
                note = Note.objects.create(eleve_id = eleve, note = colonne_note[i], resultat_id = resultat)
            i = i+1

        return redirect("schoolmanage:index_note", evaluation_id = evaluation_id)
    
    return render(request, "schoolmanage/enseignants/notes/form.html", context)